"""Capture the immutable Orion DCF baseline and execution environment.

Run from the repository root:

    python scripts/capture_baseline.py --write

`--write` is intentionally required because changing the baseline is a model
governance decision, not an incidental side effect of running the model.
"""

from __future__ import annotations

import argparse
import hashlib
import importlib.metadata
import json
import platform
import subprocess
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from orion_dcf import run_orion_dcf  # noqa: E402


EXCEL_PATH = PROJECT_ROOT / "data" / "raw" / "orion_dcf.xlsx"
SNAPSHOT_PATH = (
    PROJECT_ROOT / "artifacts" / "baseline" / "model_snapshot.json"
)
ENVIRONMENT_PATH = (
    PROJECT_ROOT / "artifacts" / "baseline" / "environment.txt"
)

BASELINE_COMMIT = "0f54ca2000c25a75649288887b806028ad5a5433"
BASELINE_ID = "orion-dcf-2025-12-31-v1"

MODEL_FILES = [
    "data/raw/orion_dcf.xlsx",
    "requirements.txt",
    "cash_flow_model.py",
    "equity_bridge.py",
    "fcff_model.py",
    "forecast_model.py",
    "orion_dcf.py",
    "valuation_model.py",
    "orion_dashboard.py",
    "orion_valuation_lab.py",
]

PACKAGE_NAMES = [
    "marimo",
    "openpyxl",
    "pandas",
    "plotly",
    "pytest",
]


def _run_git(*args: str) -> str:
    completed = subprocess.run(
        ["git", *args],
        cwd=PROJECT_ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    return completed.stdout.strip()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _iso(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _file_metadata(relative_path: str) -> dict[str, Any]:
    path = PROJECT_ROOT / relative_path
    stat = path.stat()
    return {
        "path": relative_path,
        "sha256": _sha256(path),
        "size_bytes": stat.st_size,
        "filesystem_modified_at_utc": datetime.fromtimestamp(
            stat.st_mtime,
            timezone.utc,
        ).isoformat(),
    }


def _workbook_metadata() -> dict[str, Any]:
    formula_workbook = load_workbook(
        EXCEL_PATH,
        data_only=False,
        read_only=True,
    )
    value_workbook = load_workbook(
        EXCEL_PATH,
        data_only=True,
        read_only=True,
    )

    formula_count = 0
    for worksheet in formula_workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if (
                    isinstance(cell.value, str)
                    and cell.value.startswith("=")
                ):
                    formula_count += 1

    cached_formula_errors = []
    for worksheet in value_workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if (
                    isinstance(cell.value, str)
                    and cell.value.startswith("#")
                ):
                    cached_formula_errors.append(
                        {
                            "sheet": worksheet.title,
                            "cell": cell.coordinate,
                            "value": cell.value,
                        }
                    )

    properties = formula_workbook.properties
    calculation = formula_workbook.calculation
    return {
        "document_created_at": _iso(properties.created),
        "document_modified_at": _iso(properties.modified),
        "last_modified_by": properties.lastModifiedBy,
        "sheet_count": len(formula_workbook.sheetnames),
        "sheet_names": formula_workbook.sheetnames,
        "formula_count": formula_count,
        "cached_formula_error_count": len(cached_formula_errors),
        "cached_formula_errors": cached_formula_errors,
        "full_calculation_on_load": calculation.fullCalcOnLoad,
    }


def build_snapshot() -> dict[str, Any]:
    model = run_orion_dcf(EXCEL_PATH)
    workbook = load_workbook(
        EXCEL_PATH,
        data_only=True,
        read_only=True,
    )
    overview = workbook["개요"]
    assumptions = workbook["가정"]

    model_outputs = {
        "전망": model["전망"],
        "WACC": model["WACC"],
        "DCF": model["DCF"],
        "지분가치": model["지분가치"],
    }
    output_signature = hashlib.sha256(
        json.dumps(
            model_outputs,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()

    return {
        "schema_version": "1.0.0",
        "baseline_id": BASELINE_ID,
        "captured_at_utc": datetime.now(timezone.utc).isoformat(),
        "repository": "hahnjune0118/Orion_DCF_Valuation",
        "source_branch": "main",
        "working_branch": _run_git("branch", "--show-current"),
        "source_commit": BASELINE_COMMIT,
        "valuation_date": _iso(overview["C7"].value),
        "information_date": _iso(overview["C8"].value),
        "accounting_basis": overview["C9"].value,
        "currency": "KRW",
        "model_unit": "KRW million",
        "per_share_unit": "KRW per share",
        "forecast_period": [2026, 2027, 2028, 2029, 2030],
        "discounting_convention": "year-end",
        "terminal_growth_rate": assumptions["C37"].value,
        "tolerances": {
            "rate_absolute": 1e-12,
            "amount_million_krw_absolute": 1e-6,
            "per_share_krw_absolute": 1e-6,
            "relative": 1e-12,
        },
        "input_files": {
            "data/raw/orion_dcf.xlsx": _file_metadata(
                "data/raw/orion_dcf.xlsx"
            ),
            "requirements.txt": _file_metadata("requirements.txt"),
        },
        "workbook_metadata": _workbook_metadata(),
        "output_signature_sha256": output_signature,
        "model_outputs": model_outputs,
    }


def build_environment_text(snapshot: dict[str, Any]) -> str:
    package_versions = []
    for package in PACKAGE_NAMES:
        try:
            version = importlib.metadata.version(package)
        except importlib.metadata.PackageNotFoundError:
            version = "NOT INSTALLED"
        package_versions.append(f"- {package}=={version}")

    file_lines = []
    for relative_path in MODEL_FILES:
        metadata = _file_metadata(relative_path)
        file_lines.append(
            f"- {relative_path}: sha256={metadata['sha256']}; "
            f"size={metadata['size_bytes']} bytes; "
            "filesystem_modified_at_utc="
            f"{metadata['filesystem_modified_at_utc']}"
        )

    pip_freeze = subprocess.run(
        [sys.executable, "-m", "pip", "freeze"],
        check=True,
        capture_output=True,
        text=True,
    ).stdout.strip()

    workbook = snapshot["workbook_metadata"]
    lines = [
        "ORION DCF BASELINE EXECUTION ENVIRONMENT",
        "========================================",
        f"Baseline ID: {snapshot['baseline_id']}",
        f"Captured at (UTC): {snapshot['captured_at_utc']}",
        "Initial repository state: clean main working tree",
        f"Baseline source commit: {snapshot['source_commit']}",
        f"Working branch: {snapshot['working_branch']}",
        "Repository: hahnjune0118/Orion_DCF_Valuation",
        "",
        "RUNTIME",
        "-------",
        f"Python: {platform.python_version()}",
        f"Python executable: {sys.executable}",
        f"Implementation: {platform.python_implementation()}",
        f"Platform: {platform.platform()}",
        f"Git: {_run_git('--version')}",
        "",
        "DIRECT DEPENDENCIES",
        "-------------------",
        *package_versions,
        "",
        "INPUT WORKBOOK",
        "--------------",
        f"Path: {EXCEL_PATH.relative_to(PROJECT_ROOT)}",
        "SHA-256: "
        f"{snapshot['input_files']['data/raw/orion_dcf.xlsx']['sha256']}",
        "Size: "
        f"{snapshot['input_files']['data/raw/orion_dcf.xlsx']['size_bytes']} bytes",
        "Filesystem modified at (UTC): "
        f"{snapshot['input_files']['data/raw/orion_dcf.xlsx']['filesystem_modified_at_utc']}",
        f"Workbook created at: {workbook['document_created_at']}",
        f"Workbook modified at: {workbook['document_modified_at']}",
        f"Workbook last modified by: {workbook['last_modified_by']}",
        f"Sheets ({workbook['sheet_count']}): "
        + ", ".join(workbook["sheet_names"]),
        f"Formula cells: {workbook['formula_count']}",
        f"Cached formula errors: {workbook['cached_formula_error_count']}",
        "Git history for workbook: commit "
        "e9fa4a7a00100e1a3ec078d81847c110a7a201c6 "
        "(2026-08-18T11:23:09+09:00)",
        "",
        "MODEL FILE FINGERPRINTS",
        "-----------------------",
        *file_lines,
        "",
        "VALIDATION AT PHASE 1.1",
        "-----------------------",
        "Pre-change pytest: 18 passed",
        "Post-control pytest: see docs/checkpoints/STATUS.md",
        "Marimo check: passed; container emitted only a read-only",
        "/root/.config discovery warning",
        "",
        "FULL PYTHON ENVIRONMENT (pip freeze)",
        "------------------------------------",
        pip_freeze,
        "",
    ]
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--write",
        action="store_true",
        help="Write the governed baseline files.",
    )
    args = parser.parse_args()

    snapshot = build_snapshot()
    if not args.write:
        print(json.dumps(snapshot, ensure_ascii=False, indent=2))
        return 0

    SNAPSHOT_PATH.parent.mkdir(parents=True, exist_ok=True)
    SNAPSHOT_PATH.write_text(
        json.dumps(snapshot, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    ENVIRONMENT_PATH.write_text(
        build_environment_text(snapshot),
        encoding="utf-8",
    )
    print(f"Wrote {SNAPSHOT_PATH.relative_to(PROJECT_ROOT)}")
    print(f"Wrote {ENVIRONMENT_PATH.relative_to(PROJECT_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
